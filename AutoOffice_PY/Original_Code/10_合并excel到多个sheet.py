from openpyxl import load_workbook,Workbook
import os
# 多个excel合并到1个excel中的多个sheet
def copy_data():
    wb = Workbook()
    for name in os.listdir('./销售表'):
        path = f'./销售表/{name}'
        tmp_wb = load_workbook(path)
        tmp_sh = tmp_wb.active
        sh = wb.create_sheet(name[:-5])
        for r in range(1,tmp_sh.max_row+1):
            # 获取整行数据
            row_value = []
            for c in range(1,tmp_sh.max_column+1):
                value = tmp_sh.cell(r,c).value
                row_value.append(value)
            sh.append(row_value)
    del wb['Sheet']
    wb.save('./create_data/10_合并的数据.xlsx')

if __name__ == "__main__":
    copy_data()