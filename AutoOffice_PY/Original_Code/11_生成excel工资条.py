from openpyxl import load_workbook,Workbook
from openpyxl.packaging.manifest import Manifest

def create_excel():
    wb = load_workbook('./base_data/工资数据.xlsx',data_only=True)
    sh = wb.active
    # 按照行数 max_row
    title = ['工号','姓名','部门','基本工资','提成','加班工资','社保扣除','考勤扣除','应发工资','邮箱']
    for i,row in enumerate(sh.rows):
        if i == 0:
            continue
        else:
            temp_wbook = Workbook()
            temp_sh = temp_wbook.active
            temp_sh.append(title)
            row_data = [cell.value for cell in row]
            temp_sh.append(row_data)
            print(row_data)
            temp_wbook.save(f'./create_data/{row_data[1]}.xlsx')

if __name__ == "__main__":
    create_excel()