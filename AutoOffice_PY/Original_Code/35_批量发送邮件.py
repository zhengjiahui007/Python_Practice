import smtplib
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.header import Header
smtp_obj = smtplib.SMTP('smtp.qq.com')
smtp_obj.login('398707160@qq.com','ppzuouwejpuebgfg')

# 编辑内容
wb = load_workbook('./base_data/工资数据.xlsx',data_only=True)
sh = wb.active

table_head ='<tr>'
for index , row in enumerate(sh.iter_rows()):
    if index == 0:
        for col in row:
            table_head += f'<td>{col.value}</td>'
        table_head += '</tr>'
        continue
    else:
        table_info ='<tr>'
        for col in row:
            table_info +=  f'<td>{col.value}</td>'
        table_info +='</tr>'

        name = row[1].value
        address = row[9].value

        msg_info =f'''
        <h3>您好，{name}:</h3>
        <p>请查收2030年12月的工资详情:</p>
        <table border="1">
        {table_head}
        {table_info}
        </table>
        '''

        msg_body = MIMEText(msg_info,'html','utf-8')
        msg_body['From'] = Header('人事部','utf-8')
        msg_body['Subject'] = Header('某某公司2030年12月工资条','utf-8')
        smtp_obj.sendmail('398707160@qq.com',['hotelmail@126.com'],msg_body.as_string())
        print(f'成功发送给{name}邮件 地址是：{address}')

'''
您好：{name} 请查收本月的工资详情：
<table>
    <tr>    table_head
        <td>工号</td>
        <td>姓名</td>
        <td>基本工资</td>
        <td>保险</td>
        <td>总金额</td>
    </tr>
    <tr>    table_info
        <td>1000</td>
        <td>刘备</td>
        <td>1000</td>
        <td>11</td>
        <td>9902</td>
    </tr>
</table>
'''