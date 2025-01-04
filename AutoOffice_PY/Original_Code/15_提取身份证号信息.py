from openpyxl import load_workbook
from datetime import datetime
def create_time():
    now_year = datetime.now().year
    wb = load_workbook('./base_data/person.xlsx')
    sh = wb.active
    max_column = sh.max_column
    for i,cell in enumerate(sh['B']):
        pno = cell.value
        year = pno[6:10]        # 6位行政区划 4位年 2月 2位日 4位个人识别码
        mouth = pno[10:12]
        day = pno[12:14]
        # print(f'year:{year} mouth:{mouth} day:{day}')
        age = now_year - int(year)
        sh.cell(i+1,max_column+1).value = year
        sh.cell(i+1,max_column+2).value = mouth
        sh.cell(i+1,max_column+3).value = day
        sh.cell(i+1,max_column+4).value = age
    wb.save('./create_data/15_提取身份证号信息.xlsx')
if __name__ == "__main__":
    # https://www.osgeo.cn/openpyxl/index.html
    # https://openpyxl.readthedocs.io/en/stable/
    # http://www.python-excel.org/
    create_time()