# pip install openpyxl
def open():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    # 获取工作薄
    sh1 = wb.active
    sh2 = wb['Sheet1']
    sh3 = wb.get_sheet_by_name('Sheet1')
    print(sh1 is sh2 is sh3)

def show_sheets():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    print(wb.sheetnames)
    for sh in wb:
        print(sh.title)

def get_one_value():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh1 = wb.active
    value1 = sh1.cell(2,3).value
    value2 = sh1['c2'].value
    print(value1,value2)

def get_many_value():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh1 = wb['Sheet1']
    # 切片
    cells1 = sh1['c2':'d3']
    # print(cells1)
    # 整行，整列
    cell_row3 = sh1[3]
    cell_col3 = sh1['c']
    print(cell_row3)
    print(cell_col3)
    cell_row3_5 = sh1[3:5]
    print(cell_row3_5)

    # 通过迭代获取数据
    # for row in sh1.iter_rows(min_row =2 , max_row =5, max_col =3):
    #     for cell in row:
    #         print(cell.value)

    for row in sh1.iter_rows(min_row =2 , max_row =5, min_col =2, max_col =4):
        for cell in row:
            print(cell.value)
def get_all_data():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh1 = wb['Sheet1']
    for row in sh1.rows:
        for cell in row:
            print(cell.value)

    for column in sh1.columns:
        for cell in column:
            print(cell.value)

def get_num():
    from openpyxl import load_workbook
    wb = load_workbook('./base_data/data01.xlsx')
    sh1 = wb['Sheet1']  
    print(sh1.max_row)
    print(sh1.max_column)
if __name__ == "__main__":
    # open()
    # show_sheets()
    # get_one_value()
    # get_many_value()
    # get_all_data()
    get_num()