from openpyxl import load_workbook,Workbook
import os
# 多个excel合并到1个excel中的1个sheet
def copy_data():
    wb = Workbook()
    sh = wb.active
    all_data =[]
    for name in os.listdir('./销售表'):
        path = f'./销售表/{name}'
        tmp_wb = load_workbook(path)
        tmp_sh = tmp_wb.active
        for r in range(1,tmp_sh.max_row+1):
            # 获取整行数据
            row_value = []
            for c in range(1,tmp_sh.max_column+1):
                value = tmp_sh.cell(r,c).value
                row_value.append(value)
            # 获取整行数据 加到全局数据里
            if row_value not in all_data:   # 去重操作
                all_data.append(row_value)
    for data in all_data:
        sh.append(data)
    wb.save('./create_data/09_合并的数据.xlsx')

if __name__ == "__main__":
    copy_data()