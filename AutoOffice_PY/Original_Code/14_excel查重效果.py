from openpyxl import load_workbook
from openpyxl.styles import PatternFill
def dum():
    wb = load_workbook('./base_data/打卡时间.xlsx')
    sh = wb.active
    # 存储哪一行是重复数据
    index = []
    tmp = [] # 没有重复的数据
    for i,c in enumerate(sh['B']):
        flag = c.value not in tmp
        # print(flag,f'===={c} == {tmp}')
        if flag:
            tmp.append(c.value)
        else:
            index.append(i)
    fill = PatternFill('solid',fgColor='AEEEEE')
    for i,r in enumerate(sh.rows):
        if i in index:
            for c in r:
                c.fill = fill
            print(f'第{i+1}条数据是重复数据')
    wb.save('./create_data/14_查找重复数据.xlsx')

if __name__ == "__main__":
    dum()