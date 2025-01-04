def new():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh2 = wb.create_sheet('数据')
    sh3 = wb.create_sheet('人员',0)
    wb.save('./create_data/08_openpyxl创建.xlsx')

def set_value():
    from openpyxl.styles import Font, Alignment, colors
    bold_italic_30_font = Font(name='微软雅黑',size =30, italic= True , bold= True, color =colors.BLUE)
    bold_italic_20_font = Font(name='微软雅黑',size =20, italic= True , bold= True, color ='CD1076')
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh1['A1'] ='Hello!'
    sh1['B2'] = 'Excel'
    sh1['B2'].font = bold_italic_30_font
    sh1['c3'] = 'Python'
    sh1['c3'].font = bold_italic_20_font

    wb.save('./create_data/08_openpyxl创建.xlsx')
def set_value2():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    data = ['Python','Java','C++']
    for i , d in enumerate(data):
        sh1.cell(i+1,1).value = d
    wb.save('./create_data/08_添加数据.xlsx')
def set_style2():
    from openpyxl.styles import Alignment
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh1.row_dimensions[1].height =30
    sh1.column_dimensions['A'].width =30
    data = ['Python','Java','C++']
    for i , d in enumerate(data):
        sh1.cell(i+1,1).value = d
        sh1.cell(i+1,1).alignment = Alignment(horizontal='left',vertical='bottom')
    wb.save('./create_data/08_修改样式.xlsx')

def set_merge():
    from openpyxl import Workbook
    wb = Workbook()
    sh1 = wb.active
    sh1.merge_cells('A1:c1')    # 单元格顺序一定是从小到大， 从前到后， 从上到下 
    sh1.merge_cells('d2:e5')
    sh1['a1'] = '横向合并' # 设置值时，必须是左上角的单元格
    sh1['d2'] = '多合并'
    wb.save('./create_data/08_合并单元格.xlsx')

def set_imge():
    from openpyxl import Workbook
    from datetime import date
    # 实例化
    wb = Workbook()
    # 激活 worksheet
    ws = wb.active
    rows = [
        ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
        [date(2020,12, 1), 40, 30, 25],
        [date(2020,12, 2), 40, 25, 30],
        [date(2020,12, 3), 50, 30, 45],
        [date(2020,12, 4), 30, 25, 40],
        [date(2020,12, 5), 25, 35, 30],
        [date(2020,12, 6), 20, 40, 35],
    ]
    for  row in rows:
        ws.append(row)
    from openpyxl.chart import LineChart,Reference
    c1 = LineChart()
    c1.title ='Line Chart'
    c1.x_axis.title ='Test_number'
    c1.y_axis.title = 'size'
    # min_col 从第几列开始是数据 min_row 从第几列开始读数据(第一行是title) max_row,max_col 数据取第几行和列
    data = Reference(ws,min_col=2,min_row=1,max_col=4,max_row=7)
    c1.add_data(data,titles_from_data=True)
    ws.add_chart(c1,'A9')
    wb.save('./create_data/08_设置图表.xlsx')

def set_img2():
    from openpyxl.chart import PieChart,Reference
    from openpyxl import Workbook 
    # from openpyxl.chart.series import DataPoint

    data = [
        ['名称', '数值'],
        ['苹果', 50],
        ['草莓', 30],
        ['椰子', 10],
        ['荔枝', 40],
    ]
    wb = Workbook()
    ws = wb.active

    for row in data:
        ws.append(row)

    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=5)
    data = Reference(ws, min_col=2, min_row=1, max_row=5)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Pies sold by category"

    # Cut the first slice out of the pie
    # slice = DataPoint(idx=0, explosion=20)
    # pie.series[0].data_points = [slice]

    ws.add_chart(pie, "D1")
    wb.save('./create_data/08_设置图表2.xlsx')
def set_img3():
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Series, Reference

    wb = Workbook()
    ws = wb.active

    rows = [
        ('Number', 'Batch 1', 'Batch 2'),
        (2, 10, 30),
        (3, 40, 60),
        (4, 50, 70),
        (5, 20, 10),
        (6, 10, 40),
        (7, 50, 30),
    ]

    for row in rows:
        ws.append(row)
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Bar Chart"
        chart1.y_axis.title = 'Test number'
        chart1.x_axis.title = 'Sample length (mm)'

        data = Reference(ws, min_col=2, min_row=1, max_row=7, max_col=3)
        cats = Reference(ws, min_col=1, min_row=2, max_row=7)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws.add_chart(chart1, "A10")
        wb.save('./create_data/08_设置图表3.xlsx')
if __name__ == "__main__":
    # new()
    # set_value()
    # set_value2()
    # set_style2()
    # set_merge()
    # set_imge()
    # set_img2()
    set_img3()