from openpyxl import Workbook
from datetime import date
from openpyxl.styles import PatternFill
def create_excel():
    wb = Workbook()
    sh = wb.active
    rows = [
        ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
        [date(2020,12, 1), 40, 30, 25],
        [date(2020,12, 2), 40, 25, 30],
        [date(2020,12, 3), 50, 30, 45],
        [date(2020,12, 4), 30, 25, 40],
        [date(2020,12, 5), 25, 35, 30],
        [date(2020,12, 6), 20, 40, 35],
    ]
    for  row in rows:
        sh.append(row)
    # 修改样式 - 填充背景色
    bak_color = PatternFill('solid',fgColor='AEEEEE')
    for i in range(1,sh.max_row+1):
        if i%2 == 0: # 如果行号能被2整除,就是偶数
            for cell in range(1,sh.max_column+1):
                sh.cell(i,cell).fill = bak_color
    
    wb.save('./create_data/12_隔行变色.xlsx')

if __name__ == "__main__":
    create_excel()